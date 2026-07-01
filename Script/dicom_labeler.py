#!/usr/bin/env python3
"""
DICOM label helper for converting doctor report conclusions into ML labels.

This script does not diagnose images. It links existing doctor conclusions to
DICOM studies, converts those conclusions into weak labels, and prepares review
or UniMiSSPlus downstream list files.

Typical workflow:
    python dicom_labeler.py extract LABELS.lnk DATA --output labels_raw.csv
    python dicom_labeler.py check-label-data LABELS.lnk DATA --output-dir label_data_check
    python dicom_labeler.py classify labels_raw.csv --output labels_classified.csv
    python dicom_labeler.py classify-xlsx LABELS.xlsx --output labels_all_classified.csv
    python dicom_labeler.py phrase-report labels_all_classified.csv --output-dir phrase_report
    python dicom_labeler.py build-lists labels_classified.csv DATA ../UniMiSSPlusdata --output-dir labels
    python dicom_labeler.py build-cv-lists labels_all_classified.csv DATA ../UniMiSSPlusdata --output-dir labels/vietnam_xray_cv
"""

import argparse
import csv
import hashlib
import json
import os
import random
import re
import subprocess
import sys
import unicodedata
import urllib.error
import urllib.request
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

for stream in (sys.stdout, sys.stderr):
    try:
        stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

try:
    import pydicom
except ImportError:
    print("Error: pydicom not installed. Run: pip install pydicom")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


COARSE_LABELS = ("NORMAL", "ABNORMAL", "UNCERTAIN")
DISEASE_LABELS = (
    "NORMAL",
    "INFECTION_OR_PNEUMONIA",
    "TUBERCULOSIS",
    "PLEURAL_EFFUSION",
    "PNEUMOTHORAX",
    "MASS_OR_NODULE",
    "FIBROSIS_OR_EMPHYSEMA",
    "FRACTURE",
    "OTHER_ABNORMAL",
    "UNCERTAIN",
)
NORMAL_ABNORMAL_CLASSES = {"Abnormal": 0, "Normal": 1}

LLM_PROMPT = """You convert Vietnamese chest imaging doctor conclusions into weak ML labels.

Use only the report text. Do not infer a diagnosis from outside knowledge.

Return strict JSON with these keys:
- coarse_label: one of NORMAL, ABNORMAL, UNCERTAIN
- disease_label: one of NORMAL, INFECTION_OR_PNEUMONIA, TUBERCULOSIS, PLEURAL_EFFUSION, PNEUMOTHORAX, MASS_OR_NODULE, FIBROSIS_OR_EMPHYSEMA, FRACTURE, OTHER_ABNORMAL, UNCERTAIN
- multi_labels: array using the same disease labels, excluding NORMAL unless it is the only label
- evidence: short exact phrase from the report supporting the label
- confidence: number from 0 to 1
- needs_review: true when uncertain, ambiguous, contradictory, or low confidence

Rules:
- NORMAL means the conclusion says no important chest abnormality.
- If there is any definite abnormal finding, coarse_label is ABNORMAL.
- If the report says suspected/cannot exclude/follow up/unclear, use UNCERTAIN or needs_review=true.
- The first-stage training label is only Normal versus Abnormal. Use UNCERTAIN for unclear reports.

Report conclusion:
{conclusion}
"""


def strip_accents(text: str) -> str:
    decomposed = unicodedata.normalize("NFD", text)
    return "".join(ch for ch in decomposed if unicodedata.category(ch) != "Mn")


def normalize_text(text: str) -> str:
    text = strip_accents(text or "").lower()
    text = text.replace("đ", "d")
    text = re.sub(r"[^a-z0-9.+/ -]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def contains_any(text: str, phrases) -> str:
    for phrase in phrases:
        if phrase in text:
            return phrase
    return ""


def is_negated_finding(text: str, start: int) -> bool:
    """Return True only when a finding is directly negated near its phrase."""
    before = text[max(0, start - 35):start].strip()
    negation_patterns = (
        "khong thay",
        "khong co",
        "chua thay",
        "khong phat hien",
        "khong ghi nhan",
    )
    return any(before.endswith(pattern) for pattern in negation_patterns)


def extract_conclusion(raw_text: str) -> str:
    """Extract the Vietnamese conclusion section, falling back to full text."""
    if not raw_text:
        return ""

    markers = (
        "KẾT LUẬN:",
        "KẾT LUẬN",
        "KET LUAN:",
        "KET LUAN",
        "Kết luận:",
        "kết luận:",
    )
    upper_text = raw_text.upper()
    for marker in markers:
        idx = upper_text.find(marker.upper())
        if idx != -1:
            text = raw_text[idx + len(marker):].strip()
            return text.strip("-: \n\r\t")
    return raw_text.strip()


def resolve_label_path(path: str) -> Path:
    """Resolve a direct XLSX path or a Windows .lnk shortcut to the workbook."""
    label_path = Path(path)
    if label_path.suffix.lower() != ".lnk":
        return label_path

    try:
        import win32com.client  # type: ignore

        shell = win32com.client.Dispatch("WScript.Shell")
        target = shell.CreateShortcut(str(label_path.resolve())).TargetPath
        if target:
            return Path(target)
    except Exception:
        pass

    try:
        command = (
            "$s=(New-Object -ComObject WScript.Shell).CreateShortcut("
            f"'{str(label_path.resolve())}'"
            "); $s.TargetPath"
        )
        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", command],
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )
        target = result.stdout.strip()
        if target:
            return Path(target)
    except Exception:
        pass

    return label_path


def looks_like_header(row) -> bool:
    first = str(row[0] or "").strip().lower() if row else ""
    second = str(row[1] or "").strip().lower() if len(row) > 1 else ""
    if re.fullmatch(r"[0-9.]+", first):
        return False
    header_words = ("study", "uid", "conclusion", "report", "label")
    return any(word in first for word in header_words) or any(word in second for word in header_words)


def load_labels(xlsx_path: str) -> dict:
    """Load LABELS.xlsx into {match_id: {'conclusion': str, 'notes': str}}."""
    resolved_path = resolve_label_path(xlsx_path)
    workbook = openpyxl.load_workbook(resolved_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    start_idx = 1 if rows and looks_like_header(rows[0]) else 0

    labels = {}
    for row in rows[start_idx:]:
        if not row:
            continue
        match_id = str(row[0]).strip("' ") if row[0] else ""
        conclusion = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        notes = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if match_id:
            labels[match_id] = {"conclusion": conclusion, "notes": notes}
    return labels


def iter_label_rows(xlsx_path: str) -> list[dict]:
    """Load LABELS.xlsx rows without collapsing duplicate match IDs."""
    resolved_path = resolve_label_path(xlsx_path)
    workbook = openpyxl.load_workbook(resolved_path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    start_idx = 1 if rows and looks_like_header(rows[0]) else 0

    labels = []
    for row_index, row in enumerate(rows[start_idx:], start=start_idx + 1):
        if not row:
            continue
        match_id = str(row[0]).strip("' ") if row[0] else ""
        conclusion = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        notes = str(row[2]).strip() if len(row) > 2 and row[2] else ""
        if match_id:
            labels.append({
                "xlsx_row": str(row_index),
                "label_match_id": match_id,
                "conclusion_full": conclusion,
                "conclusion": extract_conclusion(conclusion),
                "label_notes": notes,
            })
    return labels


def read_dicom(raw: bytes, stop_before_pixels: bool = False):
    buffer = BytesIO(raw)
    try:
        return pydicom.dcmread(buffer, force=False, stop_before_pixels=stop_before_pixels)
    except Exception:
        buffer.seek(0)
        return pydicom.dcmread(buffer, force=True, stop_before_pixels=stop_before_pixels)


def iter_zip_dicoms(zip_path: Path):
    with ZipFile(zip_path, "r") as archive:
        names = [name for name in archive.namelist() if name.lower().endswith(".dcm") and not name.endswith("/")]
        for name in names:
            yield name, archive.read(name)


def get_first_dicom_from_zip(zip_path: Path):
    try:
        for _, raw in iter_zip_dicoms(zip_path):
            return read_dicom(raw, stop_before_pixels=True)
    except Exception:
        return None
    return None


def get_study_uid_from_zip(zip_path: Path) -> str | None:
    ds = get_first_dicom_from_zip(zip_path)
    uid = ds.get("StudyInstanceUID", None) if ds else None
    return str(uid) if uid else None


def get_modality_from_zip(zip_path: Path) -> str:
    parent = zip_path.parent.name.lower()
    if "ct" in parent:
        return "CT"
    if "x-ray" in parent or "xray" in parent or "x_ray" in parent:
        return "X-ray"
    ds = get_first_dicom_from_zip(zip_path)
    modality = ds.get("Modality", None) if ds else None
    return str(modality) if modality else "UNKNOWN"


def classify_by_rules(conclusion: str) -> dict:
    """Rule-based weak labeler for Vietnamese chest report conclusions."""
    text = normalize_text(conclusion)
    if not text:
        return label_result(
            "NORMAL",
            "NORMAL",
            ["NORMAL"],
            "",
            0.90,
            False,
            "empty_conclusion_default_normal",
        )

    uncertain_phrase = contains_any(text, (
        "nghi", "theo doi", "chua loai tru", "khong loai tru", "kha nang",
        "co the", "can doi chieu", "de nghi", "nen chup", "chua ro",
    ))

    positive_patterns = {
        "INFECTION_OR_PNEUMONIA": (
            "viem phoi", "viem phe quan phoi", "dong dac", "tham nhiem",
            "kinh mo", "ground glass", "ggo", "nhiem trung", "mo phoi",
            "viem thuy", "viem phe quan",
        ),
        "TUBERCULOSIS": ("lao", "tuberculosis", "tb phoi"),
        "PLEURAL_EFFUSION": ("tran dich mang phoi", "dich mang phoi", "day dich mang phoi"),
        "PNEUMOTHORAX": ("tran khi khoang mang phoi", "tran khi mang phoi", "khi mang phoi"),
        "MASS_OR_NODULE": (
            "khoi", "u phoi", "not", "nodul", "nodule", "mass", "hamartoma",
            "di can", "ung thu", "carcinoma",
        ),
        "FIBROSIS_OR_EMPHYSEMA": (
            "xo phoi", "xoa phoi", "day xo", "gian phe nang", "khi phe thung",
            "emphysema", "copd", "benh phoi tac nghen", "xo kem voi",
            "xo hoa", "xo rai rac", "xoa rai rac", "xo gian phe quan",
            "gian phe quan", "dai xo", "canh xo", "ken khi",
            "ken khi thuy duoi phoi", "ken khi nho canh vach",
        ),
        "FRACTURE": (
            "gay xuong", "nut xuong", "chan thuong xuong",
            "gay 1/3 ngoai xuong don", "gay ran cung truoc xuong suon",
            "gay cung truoc xuong suon", "vo lun dot song nguc",
        ),
    }
    multi_labels = []
    evidence = []
    for label, phrases in positive_patterns.items():
        phrase = contains_any(text, phrases)
        if not phrase:
            continue
        start = text.find(phrase)
        if is_negated_finding(text, start):
            continue
        multi_labels.append(label)
        evidence.append(phrase)

    other_abnormal_phrase = contains_any(text, (
        "bat thuong", "ton thuong", "xep phoi", "phu phoi", "day thanh phe quan",
        "vong cung dong mach chu", "voi hoa", "tao hang", "u trung that",
        "hach", "day mang phoi", "dinh mang phoi", "tran dich", "ton tai",
        "mo khong thuan nhat", "tang dam", "quai dong mach chu vong",
        "voi hoa thanh quai", "thiet bi tao nhip", "cathete", "catheter",
        "day dinh", "dinh nhe mang phoi", "ron phoi",
        "cung dong mach chu vong", "dong mach chu vong", "quai dong mach chu",
        "day to chuc ke", "day ke", "dai mo", "dam mo", "mo ngoai vi",
        "mo tuong doi thuan nhat", "day dinh mang phoi", "dinh mang phoi",
        "dan luu", "dan luu khoang mang phoi", "mo goc suon hoanh",
        "goc suon hoanh ben trai tu", "goc suon hoanh trai tu",
        "xep nhe", "xep phoi", "dap phoi",
        "xuat huyet vung nhan xam", "tran mau nao that",
        "day niem mac", "tu dich da xoang", "mo nen phoi trai",
        "bong tim to", "mo thuan nhat goc suon hoanh trai",
        "quai dmc vong", "quai dong mach vong",
        "goc suon hoanh phai kem nhon", "tang cac nhanh phe huyet",
        "day cac nhanh phe huyet quan",
        "bong tim khong to truong phoi 2 ben sang",
        "mo gan hoan toan truong phoi phai",
    ))
    if other_abnormal_phrase and not multi_labels:
        start = text.find(other_abnormal_phrase)
        if not is_negated_finding(text, start):
            multi_labels.append("OTHER_ABNORMAL")
            evidence.append(other_abnormal_phrase)

    normal_phrase = contains_any(text, (
        "khong thay bat thuong",
        "khong phat hien bat thuong",
        "khong thay ton thuong",
        "khong thay hinh anh bat thuong",
        "binh thuong",
        "phoi sang deu",
        "truong phoi hai ben sang deu",
        "bong tim khong to truong phoi hai ben sang",
        "bong tim khong to hai phoi sang",
        "khong thay khoi not",
    ))

    if multi_labels:
        primary = choose_primary_label(multi_labels)
        confidence = 0.65 if uncertain_phrase else 0.9
        return label_result(
            "ABNORMAL",
            primary,
            multi_labels,
            "; ".join(evidence),
            confidence,
            bool(uncertain_phrase),
            "rule_positive" + (f":{uncertain_phrase}" if uncertain_phrase else ""),
        )

    if normal_phrase:
        confidence = 0.7 if uncertain_phrase else 0.92
        return label_result(
            "NORMAL",
            "NORMAL",
            ["NORMAL"],
            normal_phrase,
            confidence,
            bool(uncertain_phrase),
            "rule_normal" + (f":{uncertain_phrase}" if uncertain_phrase else ""),
        )

    return label_result(
        "UNCERTAIN",
        "UNCERTAIN",
        ["UNCERTAIN"],
        uncertain_phrase,
        0.35,
        True,
        "no_clear_rule_match",
    )


def choose_primary_label(labels) -> str:
    priority = (
        "TUBERCULOSIS",
        "MASS_OR_NODULE",
        "PNEUMOTHORAX",
        "PLEURAL_EFFUSION",
        "INFECTION_OR_PNEUMONIA",
        "FIBROSIS_OR_EMPHYSEMA",
        "FRACTURE",
        "OTHER_ABNORMAL",
    )
    for label in priority:
        if label in labels:
            return label
    return labels[0] if labels else "UNCERTAIN"


def normal_abnormal_mapping(coarse_label: str) -> tuple[str, str]:
    if coarse_label == "NORMAL":
        return "Normal", "1"
    if coarse_label == "ABNORMAL":
        return "Abnormal", "0"
    return "EXCLUDE", ""


def label_result(coarse, disease, multi, evidence, confidence, needs_review, reason) -> dict:
    binary_label, binary_class = normal_abnormal_mapping(coarse)
    return {
        "coarse_label": coarse,
        "disease_label": disease,
        "multi_labels": ";".join(multi),
        "evidence": evidence,
        "confidence": f"{confidence:.2f}",
        "needs_review": "yes" if needs_review or confidence < 0.75 else "no",
        "label_source": "rules",
        "label_reason": reason,
        "normal_abnormal_label": binary_label,
        "normal_abnormal_class": binary_class,
    }


def classify_with_llm(conclusion: str, api_key: str, endpoint: str, model: str) -> dict:
    payload = json.dumps({
        "model": model,
        "messages": [{"role": "user", "content": LLM_PROMPT.format(conclusion=conclusion)}],
        "temperature": 0.0,
        "max_tokens": 300,
        "response_format": {"type": "json_object"},
    }).encode("utf-8")

    request = urllib.request.Request(
        endpoint,
        data=payload,
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {api_key}",
        },
    )
    response = urllib.request.urlopen(request, timeout=60)
    body = json.loads(response.read().decode("utf-8"))
    parsed = json.loads(body["choices"][0]["message"]["content"])

    coarse = str(parsed.get("coarse_label", "UNCERTAIN")).upper()
    disease = str(parsed.get("disease_label", "UNCERTAIN")).upper()
    if coarse not in COARSE_LABELS:
        coarse = "UNCERTAIN"
    if disease not in DISEASE_LABELS:
        disease = "UNCERTAIN"

    multi = parsed.get("multi_labels", [])
    if not isinstance(multi, list):
        multi = [str(multi)]
    multi = [str(item).upper() for item in multi if str(item).upper() in DISEASE_LABELS]
    if not multi:
        multi = [disease]

    confidence = parsed.get("confidence", 0.0)
    try:
        confidence = float(confidence)
    except (TypeError, ValueError):
        confidence = 0.0
    confidence = max(0.0, min(1.0, confidence))

    result = label_result(
        coarse,
        disease,
        multi,
        str(parsed.get("evidence", "")),
        confidence,
        bool(parsed.get("needs_review", True)),
        "llm_json",
    )
    result["label_source"] = "llm"
    return result


def read_csv(path: str) -> list[dict]:
    with open(path, "r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: str | Path, rows: list[dict], preferred_fields=None):
    fields = []
    for field in preferred_fields or []:
        if field not in fields:
            fields.append(field)
    for row in rows:
        for field in row:
            if field not in fields:
                fields.append(field)

    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def cmd_extract(args):
    labels = load_labels(args.labels)
    print(f"Loaded {len(labels)} labels from {args.labels}")

    data_path = Path(args.data_dir)
    zip_files = sorted(data_path.rglob("*.zip"))
    print(f"Found {len(zip_files)} ZIP files in {args.data_dir}")

    results = []
    matched = 0
    unmatched = 0
    for zip_path in zip_files:
        uid = get_study_uid_from_zip(zip_path)
        zip_stem = zip_path.stem
        match_key = uid if uid and uid in labels else zip_stem if zip_stem in labels else None
        if match_key:
            matched += 1
            entry = labels[match_key]
            results.append({
                "study_uid": uid or "",
                "label_match_id": match_key,
                "label_match_method": "study_uid" if match_key == uid else "zip_stem",
                "modality": get_modality_from_zip(zip_path),
                "zip_path": str(zip_path.relative_to(data_path)),
                "conclusion_full": entry["conclusion"],
                "conclusion": extract_conclusion(entry["conclusion"]),
                "label_notes": entry.get("notes", ""),
            })
        else:
            unmatched += 1
            if unmatched <= 5:
                print(f"  No match: {zip_path.name} (uid={uid})")

    print(f"\nMatched: {matched}, Unmatched: {unmatched}")
    preferred = (
        "study_uid", "label_match_id", "label_match_method", "modality", "zip_path",
        "conclusion_full", "conclusion", "label_notes",
    )
    write_csv(args.output, results, preferred)
    print(f"Saved {len(results)} entries to {args.output}")


def cmd_classify(args):
    rows = read_csv(args.input)
    errors = classify_rows(rows, args)

    preferred = (
        "study_uid", "label_match_id", "label_match_method", "modality", "zip_path",
        "coarse_label", "normal_abnormal_label", "normal_abnormal_class",
        "disease_label", "multi_labels", "confidence", "needs_review", "evidence",
        "label_source", "label_reason", "conclusion", "conclusion_full", "label_notes",
    )
    write_csv(args.output, rows, preferred)
    print(f"\nSaved {len(rows)} classified entries to {args.output}")
    print_label_counts(rows)
    if errors:
        print(f"Classification errors: {errors}")


def classify_rows(rows: list[dict], args) -> int:
    api_key = args.api_key or os.environ.get("OPENAI_API_KEY", "")
    if args.method == "llm" and not api_key:
        print("Error: --method llm requires OPENAI_API_KEY or --api-key.")
        sys.exit(1)

    print(f"Classifying {len(rows)} conclusions with method={args.method}...")
    errors = 0
    for index, row in enumerate(rows):
        conclusion = row.get("conclusion", "") or row.get("conclusion_full", "")
        try:
            if args.method == "llm":
                result = classify_with_llm(conclusion, api_key, args.endpoint, args.model)
            else:
                result = classify_by_rules(conclusion)
        except (urllib.error.URLError, json.JSONDecodeError, KeyError, TimeoutError, ValueError) as exc:
            errors += 1
            result = label_result("UNCERTAIN", "UNCERTAIN", ["UNCERTAIN"], "", 0.0, True, f"error:{exc}")
        row.update(result)
        if index == 0 or (index + 1) % 10 == 0 or index == len(rows) - 1:
            print(
                f"  [{index + 1}/{len(rows)}] "
                f"{row.get('label_match_id', row.get('study_uid', ''))[:32]} -> "
                f"{row['coarse_label']}/{row['disease_label']}"
            )

    return errors


def print_label_counts(rows: list[dict]):
    print("Label counts:")
    for label, count in Counter(row["coarse_label"] for row in rows).most_common():
        print(f"  {label}: {count}")
    print("Normal/Abnormal training-label counts:")
    for label, count in Counter(row["normal_abnormal_label"] for row in rows).most_common():
        print(f"  {label}: {count}")


def cmd_classify_xlsx(args):
    rows = iter_label_rows(args.labels)
    print(f"Loaded {len(rows)} rows from {args.labels}")
    empty_count = sum(1 for row in rows if not row.get("conclusion", "").strip())
    print(f"Rows with empty conclusion defaulting to Normal: {empty_count}")

    errors = classify_rows(rows, args)
    preferred = (
        "xlsx_row", "label_match_id",
        "coarse_label", "normal_abnormal_label", "normal_abnormal_class",
        "disease_label", "multi_labels", "confidence", "needs_review", "evidence",
        "label_source", "label_reason", "conclusion", "conclusion_full", "label_notes",
    )
    write_csv(args.output, rows, preferred)
    print(f"\nSaved {len(rows)} classified Excel rows to {args.output}")
    print_label_counts(rows)
    if errors:
        print(f"Classification errors: {errors}")


def split_phrases(value: str) -> list[str]:
    phrases = []
    for part in (value or "").split(";"):
        phrase = part.strip()
        if phrase:
            phrases.append(phrase)
    return phrases


def normalized_ngrams(text: str, min_n: int = 2, max_n: int = 4) -> list[str]:
    tokens = normalize_text(text).split()
    stop_words = {
        "hinh", "anh", "x", "quang", "nguc", "thang", "phim", "ct", "scanner",
        "hai", "ben", "phoi", "truong", "va", "hoac", "la", "co", "khong",
    }
    tokens = [token for token in tokens if len(token) > 1 and token not in stop_words]
    ngrams = []
    for size in range(min_n, max_n + 1):
        for index in range(0, max(0, len(tokens) - size + 1)):
            ngrams.append(" ".join(tokens[index:index + size]))
    return ngrams


def write_counter_csv(path: Path, rows):
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(("label", "phrase", "count"))
        for label, phrase, count in rows:
            writer.writerow((label, phrase, count))


def top_by_label(rows: list[dict], key_fn, top_n: int):
    counters = defaultdict(Counter)
    for row in rows:
        label = row.get("coarse_label", "UNKNOWN") or "UNKNOWN"
        for key in key_fn(row):
            counters[label][key] += 1

    output = []
    for label in sorted(counters):
        for phrase, count in counters[label].most_common(top_n):
            output.append((label, phrase, count))
    return output


def cmd_phrase_report(args):
    rows = read_csv(args.input)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    top_n = args.top_n

    uncertain_rows = [row for row in rows if row.get("coarse_label") == "UNCERTAIN"]
    write_csv(
        output_dir / "uncertain_samples.csv",
        uncertain_rows,
        (
            "xlsx_row", "label_match_id", "label_reason", "conclusion",
            "conclusion_full", "label_notes",
        ),
    )

    write_counter_csv(
        output_dir / "common_exact_conclusions.csv",
        top_by_label(rows, lambda row: [row.get("conclusion", "").strip()] if row.get("conclusion", "").strip() else [], top_n),
    )
    write_counter_csv(
        output_dir / "common_evidence_phrases.csv",
        top_by_label(rows, lambda row: split_phrases(row.get("evidence", "")), top_n),
    )
    write_counter_csv(
        output_dir / "common_normalized_ngrams.csv",
        top_by_label(rows, lambda row: normalized_ngrams(row.get("conclusion", "")), top_n),
    )

    label_counts = Counter(row.get("coarse_label", "UNKNOWN") or "UNKNOWN" for row in rows)
    disease_counts = Counter(row.get("disease_label", "UNKNOWN") or "UNKNOWN" for row in rows)
    reason_counts = Counter(row.get("label_reason", "UNKNOWN") or "UNKNOWN" for row in rows)
    empty_default_count = reason_counts.get("empty_conclusion_default_normal", 0)

    markdown = []
    markdown.append("# Report Phrase Analysis\n")
    markdown.append("Inspired by the Vietnamese CXR report-labeling workflow in `docs/journal.pone.0276545.pdf`: normal-template filtering, keyword detection, abnormality interpolation, and manual review of unmatched text.\n")
    markdown.append("## Summary\n")
    markdown.append(f"- Total rows: {len(rows)}\n")
    markdown.append(f"- Empty conclusions defaulted to Normal: {empty_default_count}\n")
    markdown.append(f"- Uncertain rows for manual review: {len(uncertain_rows)}\n")
    markdown.append("\n## Coarse Labels\n")
    for label, count in label_counts.most_common():
        markdown.append(f"- {label}: {count}\n")
    markdown.append("\n## Disease Labels\n")
    for label, count in disease_counts.most_common():
        markdown.append(f"- {label}: {count}\n")
    markdown.append("\n## Most Common Evidence Phrases\n")
    for label, phrase, count in top_by_label(rows, lambda row: split_phrases(row.get("evidence", "")), min(top_n, 10)):
        markdown.append(f"- {label}: `{phrase}` ({count})\n")
    markdown.append("\n## Current Uncertain Samples\n")
    if uncertain_rows:
        markdown.append("| xlsx_row | label_match_id | conclusion |\n")
        markdown.append("|---:|---|---|\n")
        for row in uncertain_rows:
            conclusion = " ".join((row.get("conclusion") or "").split())
            if len(conclusion) > 180:
                conclusion = conclusion[:177] + "..."
            markdown.append(f"| {row.get('xlsx_row', '')} | {row.get('label_match_id', '')} | {conclusion} |\n")
    else:
        markdown.append("No uncertain samples.\n")

    report_path = output_dir / "phrase_report.md"
    report_path.write_text("".join(markdown), encoding="utf-8")

    print(f"Saved phrase report to {report_path}")
    print(f"Saved uncertain samples to {output_dir / 'uncertain_samples.csv'}")
    print(f"Uncertain samples: {len(uncertain_rows)}")
    print("Label counts:")
    for label, count in label_counts.most_common():
        print(f"  {label}: {count}")


def safe_export_name(ds, fallback: str) -> str:
    uid = ds.get("SOPInstanceUID", None) or ds.get("SeriesInstanceUID", None) or ds.get("StudyInstanceUID", None) or fallback
    return hashlib.sha256(str(uid).encode()).hexdigest()[:16]


def xray_png_paths_for_zip(zip_path: Path) -> list[str]:
    paths = []
    try:
        for name, raw in iter_zip_dicoms(zip_path):
            ds = read_dicom(raw, stop_before_pixels=True)
            modality = str(ds.get("Modality", "")).upper()
            body_part = str(ds.get("BodyPartExamined", "")).upper()
            if modality in ("DX", "CR", "XR", "RG") and ("CHEST" in body_part or "THORAX" in body_part):
                paths.append(f"2D_images/{safe_export_name(ds, f'{zip_path}:{name}')}.png")
    except Exception:
        return []
    return sorted(set(paths))


def collect_data_zip_rows(data_dir: str) -> list[dict]:
    """Collect ZIP file stems and relative paths from DATA."""
    root = Path(data_dir)
    if not root.exists():
        raise ValueError(f"DATA path does not exist: {data_dir}")

    if root.is_file():
        zip_paths = [root] if root.suffix.lower() == ".zip" else []
    else:
        zip_paths = sorted(path for path in root.rglob("*.zip") if path.is_file())

    rows = []
    for path in zip_paths:
        rows.append({
            "zip_stem": path.stem,
            "zip_path": path.relative_to(root).as_posix() if root.is_dir() else path.name,
            "modality_hint": path.parent.name if root.is_dir() else "",
        })
    return rows


def label_row_modality(row: dict) -> str:
    """Classify label row modality from report text convention."""
    detail = row.get("conclusion_full", "") or ""
    if not detail.strip():
        return "X-ray"
    normalized = normalize_text(detail)
    ct_markers = (
        "ct",
        "clvt",
        "cat lop vi tinh",
        "chup cat lop",
        "do day lat cat",
        "tai tao da mat phang",
        "cua so nhu mo",
        "cua so trung that",
        "tiem thuoc can quang",
    )
    return "CT" if any(marker in normalized for marker in ct_markers) else "X-ray"


def short_text(text: str, limit: int = 140) -> str:
    text = " ".join((text or "").split())
    if len(text) <= limit:
        return text
    return text[:limit - 3] + "..."


def cmd_check_label_data(args):
    label_rows = iter_label_rows(args.labels)
    if args.modality != "all":
        label_rows = [
            row for row in label_rows
            if label_row_modality(row).lower() == args.modality
        ]
    data_rows = collect_data_zip_rows(args.data_dir)
    if args.modality != "all":
        data_rows = [
            row for row in data_rows
            if row.get("modality_hint", "").lower() == args.modality
        ]

    label_counts = Counter(row["label_match_id"] for row in label_rows)
    data_counts = Counter(row["zip_stem"] for row in data_rows)

    label_rows_missing_data = [
        {**row, "label_modality": label_row_modality(row)}
        for row in label_rows
        if row["label_match_id"] not in data_counts
    ]
    missing_by_modality = Counter(row["label_modality"] for row in label_rows_missing_data)

    data_files_missing_label = [
        row for row in data_rows
        if row["zip_stem"] not in label_counts
    ]
    label_rows_by_id = defaultdict(list)
    for row in label_rows:
        label_rows_by_id[row["label_match_id"]].append(row)

    duplicate_label_ids = []
    duplicate_label_rows = []
    for label_id, rows in sorted(label_rows_by_id.items()):
        if len(rows) <= 1:
            continue
        duplicate_label_ids.append({"label_match_id": label_id, "count": str(len(rows))})
        for row in rows:
            duplicate_label_rows.append({
                "label_match_id": label_id,
                "duplicate_count": str(len(rows)),
                "xlsx_row": row.get("xlsx_row", ""),
                "label_modality": label_row_modality(row),
                "conclusion_preview": short_text(row.get("conclusion", "") or row.get("conclusion_full", "")),
            })

    duplicate_data_stems = [
        {"zip_stem": zip_stem, "count": str(count)}
        for zip_stem, count in sorted(data_counts.items())
        if count > 1
    ]

    matched_label_rows = len(label_rows) - len(label_rows_missing_data)

    print("\n" + "=" * 80)
    print("LABEL -> DATA CONSISTENCY CHECK")
    print("=" * 80)
    print(f"  Modality filter: {args.modality}")
    print(f"  Label rows: {len(label_rows)}")
    print(f"  Unique label IDs: {len(label_counts)}")
    print(f"  DATA ZIP files: {len(data_rows)}")
    print(f"  Unique DATA ZIP stems: {len(data_counts)}")
    print(f"  Label rows matched to DATA: {matched_label_rows}")
    print(f"  Label rows missing DATA ZIP: {len(label_rows_missing_data)}")
    print(f"    Missing CT label rows: {missing_by_modality.get('CT', 0)}")
    print(f"    Missing X-ray label rows: {missing_by_modality.get('X-ray', 0)}")
    print(f"  Duplicate label IDs: {len(duplicate_label_ids)}")
    print(f"  Duplicate DATA ZIP stems: {len(duplicate_data_stems)}")
    matched_data_files = len(data_rows) - len(data_files_missing_label)
    print(f"  DATA ZIP files matched to labels: {matched_data_files}")
    print(f"  DATA ZIP files missing label row: {len(data_files_missing_label)}")

    for modality in ("CT", "X-ray"):
        rows = [row for row in label_rows_missing_data if row["label_modality"] == modality]
        if rows:
            print(f"\n--- Label rows missing DATA ZIP: {modality} (first 20) ---")
            for row in rows[:20]:
                print(f"  {row.get('label_match_id', '')}")

    if data_files_missing_label:
        print("\n--- DATA ZIP files missing label row (first 20) ---")
        for row in data_files_missing_label[:20]:
            print(f"  {row.get('zip_path', '')}")

    if duplicate_label_rows:
        print("\n--- Duplicate label ID details (first 20 rows) ---")
        for row in duplicate_label_rows[:20]:
            print(
                f"  {row.get('label_match_id', '')} | "
                f"xlsx_row={row.get('xlsx_row', '')} | "
                f"{row.get('label_modality', '')} | "
                f"{row.get('conclusion_preview', '')}"
            )

    if duplicate_data_stems:
        print("\n--- Duplicate DATA ZIP stems (first 20) ---")
        for row in duplicate_data_stems[:20]:
            print(f"  {row.get('zip_stem', '')}")

    if args.output_dir:
        output_dir = Path(args.output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)
        write_csv(output_dir / "label_rows_missing_data.csv", label_rows_missing_data)
        write_csv(
            output_dir / "label_rows_missing_data_ct.csv",
            [row for row in label_rows_missing_data if row["label_modality"] == "CT"],
        )
        write_csv(
            output_dir / "label_rows_missing_data_xray.csv",
            [row for row in label_rows_missing_data if row["label_modality"] == "X-ray"],
        )
        write_csv(output_dir / "data_files_missing_label.csv", data_files_missing_label)
        write_csv(output_dir / "duplicate_label_ids.csv", duplicate_label_ids)
        write_csv(output_dir / "duplicate_label_rows.csv", duplicate_label_rows)
        write_csv(output_dir / "duplicate_data_stems.csv", duplicate_data_stems)
        write_csv(output_dir / "summary.csv", [{
            "modality_filter": args.modality,
            "label_rows": str(len(label_rows)),
            "unique_label_ids": str(len(label_counts)),
            "data_zip_files": str(len(data_rows)),
            "unique_data_zip_stems": str(len(data_counts)),
            "label_rows_matched_to_data": str(matched_label_rows),
            "label_rows_missing_data_zip": str(len(label_rows_missing_data)),
            "label_rows_missing_data_zip_ct": str(missing_by_modality.get("CT", 0)),
            "label_rows_missing_data_zip_xray": str(missing_by_modality.get("X-ray", 0)),
            "data_zip_files_missing_label_row": str(len(data_files_missing_label)),
            "duplicate_label_ids": str(len(duplicate_label_ids)),
            "duplicate_data_zip_stems": str(len(duplicate_data_stems)),
        }])
        print(f"\nCSV reports saved to: {output_dir}")

    print("=" * 80)


def split_studies(rows: list[dict], test_fraction: float, seed: int):
    by_study = defaultdict(list)
    for row in rows:
        key = row.get("label_match_id") or row.get("study_uid") or row.get("zip_path")
        by_study[key].append(row)

    keys = sorted(by_study)
    rng = random.Random(seed)
    rng.shuffle(keys)
    test_count = int(round(len(keys) * test_fraction))
    test_keys = set(keys[:test_count])

    train_rows = []
    test_rows = []
    for key, grouped_rows in by_study.items():
        if key in test_keys:
            test_rows.extend(grouped_rows)
        else:
            train_rows.extend(grouped_rows)
    return train_rows, test_rows


def write_labeled_list(path: Path, rows: list[dict]):
    lines = []
    for row in rows:
        rel_path = row["image_path"]
        label = row["normal_abnormal_class"]
        lines.append(f"{rel_path} {label}\n")
    path.write_text("".join(sorted(lines)), encoding="utf-8")


def data_zip_index(data_dir: Path) -> dict[str, list[Path]]:
    """Index DATA zip files by filename stem for LABELS.xlsx row matching."""
    index = defaultdict(list)
    if data_dir.is_file() and data_dir.suffix.lower() == ".zip":
        index[data_dir.stem].append(data_dir)
        return index
    for path in sorted(data_dir.rglob("*.zip")):
        if path.is_file():
            index[path.stem].append(path)
    return index


def preferred_xray_zip(candidates: list[Path]) -> tuple[Path | None, str]:
    """Pick a unique X-ray-looking ZIP path, or return a skip reason."""
    if not candidates:
        return None, "missing_zip"
    xray_candidates = [
        path for path in candidates
        if any(token in path.parent.name.lower() for token in ("x-ray", "xray", "x_ray"))
    ]
    if len(xray_candidates) == 1:
        return xray_candidates[0], ""
    if len(xray_candidates) > 1:
        return None, "duplicate_xray_zip_stem"
    if len(candidates) == 1:
        return candidates[0], ""
    return None, "ambiguous_zip_stem"


def resolve_label_zip_path(row: dict, data_dir: Path, index: dict[str, list[Path]]) -> tuple[Path | None, str]:
    """Resolve an extracted/classified row to a DATA zip path."""
    zip_path_text = row.get("zip_path", "")
    if zip_path_text:
        zip_path = data_dir / zip_path_text
        if zip_path.is_file():
            return zip_path, ""

    match_id = row.get("label_match_id", "") or row.get("study_uid", "")
    if not match_id:
        return None, "missing_label_match_id"
    return preferred_xray_zip(index.get(match_id, []))


def collect_normal_abnormal_xray_samples(
    rows: list[dict],
    data_dir: Path,
    unimiss_data_dir: Path,
    include_review: bool = False,
) -> tuple[list[dict], Counter]:
    """Collect exported X-ray PNG samples with Normal/Abnormal weak labels."""
    zip_index = data_zip_index(data_dir)
    samples = []
    skipped = Counter()

    for row in rows:
        if label_row_modality(row) != "X-ray":
            skipped["ct_label_row"] += 1
            continue

        label = row.get("normal_abnormal_label", "")
        class_id = row.get("normal_abnormal_class", "")
        if label == "EXCLUDE" or class_id == "":
            skipped["excluded_label"] += 1
            continue
        if row.get("needs_review") == "yes" and not include_review:
            skipped["needs_review"] += 1
            continue

        zip_path, reason = resolve_label_zip_path(row, data_dir, zip_index)
        if not zip_path:
            skipped[reason] += 1
            continue

        rel_pngs = xray_png_paths_for_zip(zip_path)
        if not rel_pngs:
            skipped["no_xray_dicoms"] += 1
            continue

        for rel_png in rel_pngs:
            if not (unimiss_data_dir / rel_png).is_file():
                skipped["missing_exported_png"] += 1
                continue
            sample = dict(row)
            sample["image_path"] = rel_png
            sample["zip_path"] = str(zip_path.relative_to(data_dir)) if zip_path.is_relative_to(data_dir) else str(zip_path)
            sample["cv_group"] = sample.get("label_match_id") or sample.get("study_uid") or zip_path.stem
            samples.append(sample)

    return samples, skipped


def class_counts(rows: list[dict]) -> Counter:
    return Counter(row.get("normal_abnormal_label", "") for row in rows)


def grouped_samples(samples: list[dict]) -> list[dict]:
    by_group = defaultdict(list)
    for sample in samples:
        by_group[sample["cv_group"]].append(sample)

    groups = []
    for key, rows in by_group.items():
        label_counts = Counter(row["normal_abnormal_class"] for row in rows)
        label = sorted(label_counts.items(), key=lambda item: (-item[1], item[0]))[0][0]
        groups.append({
            "key": key,
            "rows": rows,
            "label": label,
            "size": len(rows),
            "mixed_labels": "yes" if len(label_counts) > 1 else "no",
        })
    return groups


def stratified_group_folds(samples: list[dict], folds: int, seed: int) -> list[list[dict]]:
    if folds < 2:
        raise ValueError("--folds must be at least 2")

    groups_by_label = defaultdict(list)
    rng = random.Random(seed)
    for group in grouped_samples(samples):
        groups_by_label[group["label"]].append(group)

    fold_groups = [[] for _ in range(folds)]
    fold_class_counts = [Counter() for _ in range(folds)]
    fold_total_counts = [0 for _ in range(folds)]

    for label, groups in sorted(groups_by_label.items()):
        rng.shuffle(groups)
        groups.sort(key=lambda group: (-group["size"], group["key"]))
        for group in groups:
            target = min(
                range(folds),
                key=lambda idx: (fold_class_counts[idx][label], fold_total_counts[idx], idx),
            )
            fold_groups[target].append(group)
            fold_class_counts[target][label] += group["size"]
            fold_total_counts[target] += group["size"]

    return [[row for group in groups for row in group["rows"]] for groups in fold_groups]


def stratified_group_validation_split(train_pool: list[dict], val_fraction: float, seed: int) -> tuple[list[dict], list[dict]]:
    if not 0.0 < val_fraction < 1.0:
        raise ValueError("--val-fraction must be between 0 and 1")

    groups_by_label = defaultdict(list)
    rng = random.Random(seed)
    for group in grouped_samples(train_pool):
        groups_by_label[group["label"]].append(group)

    val_group_keys = set()
    for label, groups in groups_by_label.items():
        rng.shuffle(groups)
        total = sum(group["size"] for group in groups)
        target = max(1, int(round(total * val_fraction))) if total else 0
        selected = 0
        for group in sorted(groups, key=lambda group: (group["size"], group["key"])):
            if selected >= target:
                break
            val_group_keys.add(group["key"])
            selected += group["size"]

    train_rows = []
    val_rows = []
    for group in grouped_samples(train_pool):
        if group["key"] in val_group_keys:
            val_rows.extend(group["rows"])
        else:
            train_rows.extend(group["rows"])
    return train_rows, val_rows


def oversample_abnormal(rows: list[dict], repeat: int) -> list[dict]:
    if repeat < 1:
        raise ValueError("--abnormal-repeat must be at least 1")
    output = []
    for row in rows:
        times = repeat if row.get("normal_abnormal_class") == "0" else 1
        output.extend([row] * times)
    return output


def write_split_summary(path: Path, fold_rows: list[dict]):
    preferred = (
        "fold", "split", "samples", "groups", "normal", "abnormal",
        "class_0_abnormal", "class_1_normal", "mixed_label_groups",
    )
    write_csv(path, fold_rows, preferred)


def split_summary_row(fold: int, split: str, rows: list[dict]) -> dict:
    groups = grouped_samples(rows)
    counts = Counter(row.get("normal_abnormal_class", "") for row in rows)
    label_counts = class_counts(rows)
    return {
        "fold": str(fold),
        "split": split,
        "samples": str(len(rows)),
        "groups": str(len(groups)),
        "normal": str(label_counts.get("Normal", 0)),
        "abnormal": str(label_counts.get("Abnormal", 0)),
        "class_0_abnormal": str(counts.get("0", 0)),
        "class_1_normal": str(counts.get("1", 0)),
        "mixed_label_groups": str(sum(1 for group in groups if group["mixed_labels"] == "yes")),
    }


def cmd_build_lists(args):
    rows = read_csv(args.input)
    data_dir = Path(args.data_dir)
    unimiss_data_dir = Path(args.unimissplus_data_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    samples, skipped = collect_normal_abnormal_xray_samples(
        rows,
        data_dir,
        unimiss_data_dir,
        include_review=args.include_review,
    )

    train_rows, test_rows = split_studies(samples, args.test_fraction, args.seed)
    train_path = output_dir / "normal_abnormal_train.txt"
    test_path = output_dir / "normal_abnormal_test.txt"
    manifest_path = output_dir / "normal_abnormal_manifest.csv"

    write_labeled_list(train_path, train_rows)
    write_labeled_list(test_path, test_rows)
    preferred = (
        "image_path", "normal_abnormal_label", "normal_abnormal_class",
        "label_match_id", "modality", "zip_path", "coarse_label", "disease_label",
        "confidence", "needs_review", "evidence", "conclusion",
    )
    write_csv(manifest_path, samples, preferred)

    print(f"Saved {len(train_rows)} train samples to {train_path}")
    print(f"Saved {len(test_rows)} test samples to {test_path}")
    print(f"Saved manifest with {len(samples)} samples to {manifest_path}")
    print("Sample class counts:")
    for label, count in Counter(row["normal_abnormal_label"] for row in samples).most_common():
        print(f"  {label}: {count}")
    if skipped:
        print("Skipped:")
        for reason, count in skipped.most_common():
            print(f"  {reason}: {count}")


def cmd_build_cv_lists(args):
    rows = read_csv(args.input)
    data_dir = Path(args.data_dir)
    unimiss_data_dir = Path(args.unimissplus_data_dir)
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    samples, skipped = collect_normal_abnormal_xray_samples(
        rows,
        data_dir,
        unimiss_data_dir,
        include_review=args.include_review,
    )
    if not samples:
        raise ValueError("No eligible exported X-ray samples found. Check DATA path and UniMiSSPlus export path.")

    folds = stratified_group_folds(samples, args.folds, args.seed)
    summary_rows = []
    manifest_rows = []
    preferred_manifest = (
        "fold", "split", "image_path", "normal_abnormal_label", "normal_abnormal_class",
        "label_match_id", "cv_group", "zip_path", "coarse_label", "disease_label",
        "confidence", "needs_review", "evidence", "conclusion",
    )

    for fold_idx in range(args.folds):
        fold_dir = output_dir / f"fold_{fold_idx}"
        fold_dir.mkdir(parents=True, exist_ok=True)

        test_rows = folds[fold_idx]
        train_pool = [row for idx, fold_rows in enumerate(folds) if idx != fold_idx for row in fold_rows]
        train_rows, val_rows = stratified_group_validation_split(
            train_pool,
            args.val_fraction,
            args.seed + fold_idx + 1,
        )
        train_oversampled_rows = oversample_abnormal(train_rows, args.abnormal_repeat)

        write_labeled_list(fold_dir / "train.txt", train_rows)
        write_labeled_list(fold_dir / "train_oversampled.txt", train_oversampled_rows)
        write_labeled_list(fold_dir / "val.txt", val_rows)
        write_labeled_list(fold_dir / "test.txt", test_rows)

        summary_rows.extend([
            split_summary_row(fold_idx, "train", train_rows),
            split_summary_row(fold_idx, "train_oversampled", train_oversampled_rows),
            split_summary_row(fold_idx, "val", val_rows),
            split_summary_row(fold_idx, "test", test_rows),
        ])

        for split_name, split_rows in (
            ("train", train_rows),
            ("val", val_rows),
            ("test", test_rows),
        ):
            for row in split_rows:
                manifest_row = dict(row)
                manifest_row["fold"] = str(fold_idx)
                manifest_row["split"] = split_name
                manifest_rows.append(manifest_row)

    write_split_summary(output_dir / "cv_split_summary.csv", summary_rows)
    write_csv(output_dir / "cv_manifest.csv", manifest_rows, preferred_manifest)

    print(f"Saved {args.folds}-fold CV lists to {output_dir}")
    print(f"Eligible samples: {len(samples)}")
    print("Eligible class counts:")
    for label, count in class_counts(samples).most_common():
        print(f"  {label}: {count}")
    print("Split summary saved to:")
    print(f"  {output_dir / 'cv_split_summary.csv'}")
    print(f"Manifest saved to:")
    print(f"  {output_dir / 'cv_manifest.csv'}")
    if skipped:
        print("Skipped:")
        for reason, count in skipped.most_common():
            print(f"  {reason}: {count}")


def main():
    parser = argparse.ArgumentParser(description="DICOM doctor-report label helper")
    subparsers = parser.add_subparsers(dest="command", help="Command")

    extract_parser = subparsers.add_parser("extract", help="Match ZIPs to LABELS.xlsx and extract conclusions")
    extract_parser.add_argument("labels", help="Path to LABELS.xlsx or LABELS.lnk")
    extract_parser.add_argument("data_dir", help="Path to DATA directory containing ZIP files")
    extract_parser.add_argument("--output", "-o", default="labels_raw.csv", help="Output CSV path")

    check_parser = subparsers.add_parser("check-label-data", help="Compare LABELS.xlsx IDs with DATA ZIP filenames")
    check_parser.add_argument("labels", help="Path to LABELS.xlsx or LABELS.lnk")
    check_parser.add_argument("data_dir", help="Path to DATA directory containing ZIP files")
    check_parser.add_argument("--output-dir", help="Directory for CSV mismatch reports")
    check_parser.add_argument(
        "--modality",
        choices=("all", "x-ray", "ct"),
        default="all",
        help="Check only labels/DATA files for one modality",
    )
    classify_parser = subparsers.add_parser("classify", help="Convert conclusions to weak labels")
    classify_parser.add_argument("input", help="Input CSV from extract")
    classify_parser.add_argument("--output", "-o", default="labels_classified.csv", help="Output CSV path")
    classify_parser.add_argument("--method", choices=("rules", "llm"), default="rules", help="Classification method")
    classify_parser.add_argument("--api-key", help="OpenAI API key for --method llm")
    classify_parser.add_argument(
        "--endpoint",
        default="https://api.openai.com/v1/chat/completions",
        help="OpenAI-compatible chat completions endpoint",
    )
    classify_parser.add_argument("--model", default="gpt-4o-mini", help="Model name for --method llm")

    xlsx_parser = subparsers.add_parser("classify-xlsx", help="Classify every LABELS.xlsx row without DICOM matching")
    xlsx_parser.add_argument("labels", help="Path to LABELS.xlsx or LABELS.lnk")
    xlsx_parser.add_argument("--output", "-o", default="labels_all_classified.csv", help="Output CSV path")
    xlsx_parser.add_argument("--method", choices=("rules", "llm"), default="rules", help="Classification method")
    xlsx_parser.add_argument("--api-key", help="OpenAI API key for --method llm")
    xlsx_parser.add_argument(
        "--endpoint",
        default="https://api.openai.com/v1/chat/completions",
        help="OpenAI-compatible chat completions endpoint",
    )
    xlsx_parser.add_argument("--model", default="gpt-4o-mini", help="Model name for --method llm")

    phrase_parser = subparsers.add_parser("phrase-report", help="Summarize common report phrases and uncertain rows")
    phrase_parser.add_argument("input", help="Classified CSV from classify or classify-xlsx")
    phrase_parser.add_argument("--output-dir", default="phrase_report", help="Directory for phrase report outputs")
    phrase_parser.add_argument("--top-n", type=int, default=30, help="Top phrases to write per label")

    lists_parser = subparsers.add_parser("build-lists", help="Build normal/abnormal fixed split lists")
    lists_parser.add_argument("input", help="Classified CSV from classify")
    lists_parser.add_argument("data_dir", help="Original DATA directory containing ZIP files")
    lists_parser.add_argument("unimissplus_data_dir", help="Directory containing exported 2D_images/")
    lists_parser.add_argument("--output-dir", default="labels", help="Directory for generated list files")
    lists_parser.add_argument("--test-fraction", type=float, default=0.2, help="Study-level test split fraction")
    lists_parser.add_argument("--seed", type=int, default=1234, help="Study-level split seed")
    lists_parser.add_argument("--include-review", action="store_true", help="Include rows flagged needs_review=yes")

    cv_parser = subparsers.add_parser(
        "build-cv-lists",
        help="Build stratified grouped k-fold normal/abnormal lists",
    )
    cv_parser.add_argument("input", help="Classified CSV from classify or classify-xlsx")
    cv_parser.add_argument("data_dir", help="Original DATA directory containing ZIP files")
    cv_parser.add_argument("unimissplus_data_dir", help="Directory containing exported 2D_images/")
    cv_parser.add_argument(
        "--output-dir",
        default="labels/vietnam_xray_cv",
        help="Directory for generated fold list files",
    )
    cv_parser.add_argument("--folds", type=int, default=5, help="Number of CV folds")
    cv_parser.add_argument(
        "--val-fraction",
        type=float,
        default=0.15,
        help="Validation fraction taken from each non-test training pool",
    )
    cv_parser.add_argument(
        "--abnormal-repeat",
        type=int,
        default=3,
        help="Total repetitions for Abnormal rows in train_oversampled.txt",
    )
    cv_parser.add_argument("--seed", type=int, default=2026, help="Grouped CV split seed")
    cv_parser.add_argument("--include-review", action="store_true", help="Include rows flagged needs_review=yes")

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        return

    if args.command == "extract":
        cmd_extract(args)
    elif args.command == "check-label-data":
        cmd_check_label_data(args)
    elif args.command == "classify":
        cmd_classify(args)
    elif args.command == "classify-xlsx":
        cmd_classify_xlsx(args)
    elif args.command == "phrase-report":
        cmd_phrase_report(args)
    elif args.command == "build-lists":
        cmd_build_lists(args)
    elif args.command == "build-cv-lists":
        cmd_build_cv_lists(args)


if __name__ == "__main__":
    main()
